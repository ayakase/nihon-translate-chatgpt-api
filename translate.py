import openpyxl
import openai
openai.api_key = ""
input_file = './data.xlsx'
workbook = openpyxl.load_workbook(input_file)

sheet_names = workbook.sheetnames
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and not cell.value.isascii() and '=' not in cell.value:
                print (cell.value)
                response = openai.ChatCompletion.create(
                    model="gpt-3.5-turbo",
                    messages=[{"role": "user", "content": ("translate this japanese to english : " + cell.value ),},
                ]
            )
                modified_content = response['choices'][0]['message']['content']
                cell.value = modified_content
            elif isinstance(cell.value, int):
                continue
            else:
                continue
output_file = 'output.xlsx'
workbook.save(output_file)